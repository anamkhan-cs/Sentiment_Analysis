from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import os
import re
import openpyxl as op
import time

analyser = SentimentIntensityAnalyzer()

start = time.time()
def sentiment_analyzer_scores(sentence):
    score = analyser.polarity_scores(sentence)

    print("{:-<40} {}".format(sentence, str(score)))
    a = re.findall("[-+]?[.]?[\d]+(?:,\d\d\d)*[\.]?\d*(?:[eE][-+]?\d+)?", str(score))#scores in string
    #writing values to cells
    neg = sheet.cell(i, 27)
    neg.value = a[0]
    neg = sheet.cell(i, 28)
    neg.value = a[1]
    neg = sheet.cell(i, 29)
    neg.value = a[2]
    neg = sheet.cell(i, 30)
    neg.value = a[3]


def sentiment(inp):
    sentiment_analyzer_scores(inp)


# with open('testing.csv', 'r') as csvfile:   #importing CSV input
#     spamreader = csv.reader(csvfile, delimiter=',')
#     for row in spamreader:
#         var1 = (','.join(row))
#         tweet_list = (var1.split(","))
#         sentiment((tweet_list[0]))

folder = "CSV"
all_files = os.listdir(folder)


for f in all_files:
    if f.endswith(".xlsx"):
        excel = op.load_workbook("CSV/"+f)

        sheet = excel["Stream"]
        row_count = sheet.max_row
        Head = sheet.cell(1, 27)
        Head.value = "Negative"
        Head = sheet.cell(1, 28)
        Head.value = "Neutral"
        Head = sheet.cell(1, 29)
        Head.value = "Positive"
        Head = sheet.cell(1, 30)
        Head.value = "Compound"

        for i in range(2, row_count):
            cell = sheet.cell(i, 7)
            # pnirint(cell.value)
            sentiment(cell.value)

        excel.save("CSV/Analysed_"+f)


end =time.time()
print("Time taken " + str(end-start))
