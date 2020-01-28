import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook
import os
d=[]
for file in os.listdir("new"):
    if file.endswith(".xls"):

		xlsBook = xlrd.open_workbook(os.path.join("new", file))
		workbook = openpyxlWorkbook()

		for i in xrange(0, xlsBook.nsheets):
		    xlsSheet = xlsBook.sheet_by_index(i)
		    sheet = workbook.active if i == 0 else workbook.create_sheet()
		    sheet.title = xlsSheet.name

		    for row in xrange(0, xlsSheet.nrows):
		        for col in xrange(0, xlsSheet.ncols):
		            sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

		workbook.save(os.path.join("new", file+'x'))

# The new xlsx file is in "workbook", without iterators (iter_rows).
# For iteration, use "for row in worksheet.rows:".
# For range iteration, use "for row in worksheet.range("{}:{}".format(startCell, endCell)):".